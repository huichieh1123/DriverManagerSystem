from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from io import BytesIO
from typing import List, Optional

from app.api.v1.schemas.jobs import Job, JobCreate, JobStatus
from app.api.v1.schemas.users import User, RoleType, DispatcherAssociationStatus
from app.api.v1.schemas.invitations import Invitation, InvitationCreate, InvitationStatus
from app.crud import job, invitation, user
from app.api.v1.endpoints.users import get_current_user # Keep this for now, will refactor users.py later

router = APIRouter()

# Dependency to check if the current user is a dispatcher
async def get_current_dispatcher(current_user: User = Depends(get_current_user)) -> User:
    if RoleType.DISPATCHER.value not in current_user.roles:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only dispatchers can perform this action")
    return current_user

# Dependency to check if the current user is a dispatcher or company
async def get_current_dispatcher_or_company(current_user: User = Depends(get_current_user)) -> User:
    if RoleType.DISPATCHER.value not in current_user.roles and RoleType.COMPANY.value not in current_user.roles:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only dispatchers or companies can perform this action")
    return current_user

@router.post("/jobs/", response_model=Job, status_code=status.HTTP_201_CREATED)
async def create_new_job(
    job_in: JobCreate,
    current_dispatcher: User = Depends(get_current_dispatcher)
):
    print(f"[dispatchers.py] Creating job for dispatcher ID: {current_dispatcher.id}")
    print(f"[dispatchers.py] Dispatcher company_id: {current_dispatcher.company_id}, company_name: {current_dispatcher.company_name}")
    new_job = await job.create(job_in, current_dispatcher.id, current_dispatcher.company_id, current_dispatcher.company_name)
    return new_job

@router.get("/jobs/template", summary="Download Job Upload Template (Excel)")
async def download_job_template(
    current_user: User = Depends(get_current_dispatcher_or_company)
):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Template"

    # Define headers for the Excel file
    headers = [
        "company", "transfer_type", "pick_up_date", "pick_up_time", "flight_number",
        "passenger_name", "phone_number", "vehicle_model", "num_of_passenger",
        "from_location", "to_location", "additional_services", "special_requirements",
        "other_contact_info", "order_number", "total_price", "email", "driver_name",
        "driver_phone", "vehicle_number", "vehicle_type", "is_public", "status"
    ]
    ws.append(headers)

    # Add some example data (optional)
    ws.append([
        "Example Company", "Airport Transfer", "2024-12-25", "14:30", "BR123",
        "John Doe", "0912345678", "Sedan", "2", "Taoyuan Airport", "Taipei City",
        "Extra luggage", "Child seat needed", "", "ORD12345", "1500",
        "john.doe@example.com", "Driver Mike", "0987654321", "ABC-1234", "Sedan",
        "TRUE", "pending"
    ])

    # Save the workbook to a BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Return the Excel file as a StreamingResponse
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=job_template.xlsx"
        }
    )

@router.post("/jobs/upload", summary="Upload Jobs from Excel File")
async def upload_jobs_from_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_dispatcher_or_company)
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid file format. Only .xlsx or .xls files are allowed.")

    try:
        # Read the Excel file
        workbook = load_workbook(filename=BytesIO(await file.read()))
        sheet = workbook.active

        # Assuming the first row is headers
        headers = [cell.value for cell in sheet[1]]
        expected_headers = [
            "company", "transfer_type", "pick_up_date", "pick_up_time", "flight_number",
            "passenger_name", "phone_number", "vehicle_model", "num_of_passenger",
            "from_location", "to_location", "additional_services", "special_requirements",
            "other_contact_info", "order_number", "total_price", "email", "driver_name",
            "driver_phone", "vehicle_number", "vehicle_type", "is_public", "status"
        ]

        if not all(h in headers for h in expected_headers):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Missing required headers in Excel file. Expected: {expected_headers}")

        created_jobs = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if not any(row): # Skip empty rows
                continue
            
            row_data = dict(zip(headers, row))
            
            try:
                # Validate and create JobCreate object
                job_data = {k: row_data.get(k) for k in expected_headers}
                job_data["is_public"] = str(job_data.get("is_public", "FALSE")).upper() == "TRUE"
                job_data["status"] = job_data.get("status", JobStatus.PENDING.value)

                # Ensure company_id and created_by_dispatcher_id are set from current_user
                job_data["company_id"] = current_user.company_id
                job_data["created_by_dispatcher_id"] = current_user.id
                job_data["company_name"] = current_user.company_name

                # Validate status
                if job_data["status"] not in [s.value for s in JobStatus]:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Invalid status '{job_data["status"]}' at row {row_index + 1}. Allowed: {', '.join([s.value for s in JobStatus])}")

                job_create = JobCreate(**job_data)
                
                # Create job using CRUD operation
                new_job = await job.create(job_create, current_user.id, current_user.company_id, current_user.company_name)
                created_jobs.append(new_job)
            except Exception as e:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Error processing row {row_index + 1}: {e}")

        return {"message": f"Successfully uploaded and created {len(created_jobs)} jobs.", "created_jobs": created_jobs}

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed to process Excel file: {e}")

@router.get("/invitations", response_model=List[Invitation])
async def get_my_invitations(
    current_dispatcher: User = Depends(get_current_dispatcher)
):
    return await invitation.get_for_invitee(current_dispatcher.id, RoleType.DISPATCHER)

@router.put("/invitations/{invitation_id}/accept", response_model=Invitation)
async def accept_invitation(
    invitation_id: str, # Changed from int to str
    current_dispatcher: User = Depends(get_current_dispatcher)
):
    inv = await invitation.get_by_id(invitation_id)
    if not inv:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Invitation not found")

    if inv["invitee_id"] != current_dispatcher.id or inv["invitee_role"] != RoleType.DISPATCHER.value:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized to accept this invitation or invitation is not for a dispatcher")

    if inv["status"] != InvitationStatus.PENDING.value:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invitation is not pending")

    # Update invitation status
    updated_inv = await invitation.update(invitation_id, {"status": InvitationStatus.ACCEPTED.value})

    # Update dispatcher's company_id, company_name, and association status
    await user.update(current_dispatcher.id, {
        "company_id": inv["company_id"],
        "company_name": inv["company_name"],
        "dispatcher_association_status": DispatcherAssociationStatus.ASSOCIATED
    })

    return updated_inv

@router.put("/invitations/{invitation_id}/decline", response_model=Invitation)
async def decline_invitation(
    invitation_id: str, # Changed from int to str
    current_dispatcher: User = Depends(get_current_dispatcher)
):
    inv = await invitation.get_by_id(invitation_id)
    if not inv:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Invitation not found")

    if inv["invitee_id"] != current_dispatcher.id or inv["invitee_role"] != RoleType.DISPATCHER.value:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized to decline this invitation or invitation is not for a dispatcher")

    if inv["status"] != InvitationStatus.PENDING.value:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invitation is not pending")

    # Update invitation status
    updated_inv = await invitation.update(invitation_id, {"status": InvitationStatus.DECLINED.value})

    # Update dispatcher's association status to UNASSOCIATED if it was PENDING
    if current_dispatcher.dispatcher_association_status == DispatcherAssociationStatus.PENDING.value:
        await user.update(current_dispatcher.id, {"dispatcher_association_status": DispatcherAssociationStatus.UNASSOCIATED})

    return updated_inv